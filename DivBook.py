"""
Create a Workbook class
"""
from asyncore import read
import imp
from statistics import mode
from tokenize import maybe
import xlsxwriter as Workbook
from typing import List
import PyStock
import pandas as pd
import os
from openpyxl import load_workbook


class DivBook:

    bookName = ""
    main_stock_content = {}
    mainSheet = "RoadTo10"

    def __init__(self, name, reader = False):
        # if not reader:
        #     self.ewriter = pd.ExcelWriter(name + ".xlsx", engine='xlsxwriter')
        #     self.book = self.ewriter.book
        #     print(self.book)
        # else:
        #     self.ereader = pd.read_excel(name + ".xlsx")
        self.bookName = name + ".xlsx"

    def addStockToMain(self, ):
        # Check if the WB is open, if open close it 
        # Close WB
        self.saveBook()

        # Get last row within sheet
        try:
            last_row = self.get_last_row(self.mainSheet)
        except KeyError:
            print("need to make sheet")
            self.addSheet(self.mainSheet)

        # # Open back up sheet 
        # print(last_row)

        # #open back up book
        # self.open_book(reader = False)
        # # print(mybook.mainSheet)
        # sheet = self.book.get_worksheet_by_name(self.mainSheet)
        # print(sheet)
        # self.addData(sheet, [(last_row, 0, "hello")])



    def open_book(self, reader = False):
        if not reader:
            self.ewriter = pd.ExcelWriter(self.bookName, engine='xlsxwriter')
            self.book = self.ewriter.book
            print(self.book)
        else:
            self.ereader = pd.read_excel(self.bookName)


    def getMainStockContent(self):
        # Get the main sheet 
        sheet = self.getSheet(self.mainSheet)

        contents = pd.read_excel(self.book, engine="xlrd")
        print(contents)


    def getSheet(self, sheetName):
        return self.book.get_worksheet_by_name(sheetName)

    """
    setActiveSheet
    """

    def setActiveSheet(self, sheet):
        return self.book.set_active_sheet(sheet)

    """
    addData
    """

    def addData(self, sheet, rows):
        # if type(rows) != List:
        #     rows = [rows]

        for row in rows:
            sheet.write(row[0], row[1], row[2])

    """
    addSheet 
      @param sheetTitle: title of sheet to be added
    """

    def addSheet(self, sheetTitle):
        sheet = self.book.add_worksheet(sheetTitle)

    """
    saveBook
      @param fileName: fileName to save book as 
    """

    def saveBook(self):
        return self.book.close()

    """
    addMainHeader
    """

    def addMainHeaderSheet(self):
        #open book
        self.open_book(reader=False)

        self.addSheet(self.mainSheet)
        sheet = self.book.get_worksheet_by_name(self.mainSheet)

        headerTitle = [
            "Symbol", "Company Name", "% Dividend Yield", "Current Price",
            "Payout Per Stock", "", "Road to 10", "Shares #", "Amount $",
            "Dividend Return $", "Rule 72"
        ]

        for i, title in enumerate(headerTitle):
            print("Prining out {}".format(title))
            self.addData(sheet, [(0, i, title)])

        self.saveBook()
    """
    TODO: 
      Push code home from home to main branch
      Clean up 
      Branch out!!!
    
    
      Create page for specific stock
        SheetName = Stock Symbol - Company Name (O - Reality Income)
    
      Track dividend buys/earnings 
        Transaction
          Date 
          Price Bought At 
          Current Price 
          * Gain 
          Div Yield
          Div Amount $
    
          Total Row
      
    """
    """
      Function to add a stock sheet to workbook
      @param stockSymbol: Stock to create sheet for
    """

    def addStockSheet(self, stockSymbol):
        #open book
        self.open_book(reader=False)
        # Create a sheet with the stock symbol being the name
        sheetName = stockSymbol
        # self.addSheet(sheetName)
        self.addSheet(sheetName)
        sheet = self.book.get_worksheet_by_name(sheetName)

        #Header for sheet
        # stockHeader = ["Transaction Date", "Price Bought At",
        #               "# Shares", "Dividend Yield %", "Dividend Amount"]

        stockHeader = [
            "Transaction Date", "Amount Sent", "Dividend Share Cost", "Shares",
            "Shares via Div", "Shares Outright"
        ]

        for i, title in enumerate(stockHeader):
            self.addData(sheet, [(0, i, title)])

        self.addData(sheet, [(1, 0, "Total: ")])

        stock = PyStock.PyStock(stockSymbol)
        stock.getMainInfo()

    def addStock(self, stockSymbol):
        """
        To add stock to main sheet 
            - read in all the contents of the main sheet into a dict
            - append new content 
        """
        #Get summary sheet
        # sheet = self.book.get_worksheet_by_name("RoadTo10")

        #Get info on stock

        stock = PyStock.PyStock(stockSymbol)
        stock.getMainInfo()
        # Sym, Company Name, divdend %
        # Get next line
        


        # self.read_sheet()

    """
    """

    def stockTotalRow(self):
        pass

    def read_sheet(self, sheetName):
        dataframe1 = pd.read_excel('DividendWorkBook.xlsx', engine="openpyxl")


        print(dataframe1)
        # print(dataframe1.sheetnames)

    
    def get_last_row(self, sheetName):
        # make sure sheet exists 

        #open 
        df = load_workbook(self.bookName) # dataframe
        sheet = df.get_sheet_by_name(sheetName)

        last_row = 0 
        for i in sheet.iter_rows():
            last_row+=1

        return last_row
